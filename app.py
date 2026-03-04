

from flask import Flask, render_template, Response, request, jsonify, send_file, redirect, url_for, session
from werkzeug.security import generate_password_hash, check_password_hash
import cv2
import numpy as np
import time
import threading
from datetime import datetime, timedelta, date
import io
import json
import os
import subprocess
import sys
import gc
import sqlite3
import atexit

# System stats for terminal (CPU, GPU, load)
try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False

# Import libraries for Excel and PDF generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel download will not work.")

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("Warning: reportlab not installed. PDF download will not work.")

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "silvermill-secret-key-change-in-production")

# Default admin credentials (used only to seed DB if no users exist)
DEFAULT_ADMIN_USERNAME = "admin"
DEFAULT_ADMIN_PASSWORD = "admin@123"

# ---------------- App start time and system events (for dashboard) ----------------
_app_started_at = datetime.now()
_system_events = []
_system_events_lock = threading.Lock()
SYSTEM_EVENTS_MAX = 200


def _log_system_event(message, level="info", event_type=None):
    """Append a system event for dashboard and persist to DB for history."""
    now = datetime.now()
    iso = now.isoformat()
    if event_type is None:

        event_type = "general"
    with _system_events_lock:
        _system_events.append({
            "time": iso,
            "message": message,
            "level": level,
            "event_type": event_type,
        })
        while len(_system_events) > SYSTEM_EVENTS_MAX:
            _system_events.pop(0)
    try:
        _insert_system_event_to_db(message=message, level=level, event_type=event_type, created_at=now)
    except Exception as e:
        print(f"[system_events_history] insert warning: {e}")


# ---------------- SQLite database for coconut count history ----------------
# Persists data from coconut_count page; weight_check page fetches from DB
_db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "instance", "silvermill.db")
_db_lock = threading.Lock()


def _get_db_conn():
    """Return a thread-local-style connection; caller must not hold _db_lock for long."""
    os.makedirs(os.path.dirname(_db_path), exist_ok=True)
    conn = sqlite3.connect(_db_path, timeout=15.0)
    conn.row_factory = sqlite3.Row
    return conn


def _init_coconut_db():
    """Create coconut_count_history table if it does not exist."""
    conn = _get_db_conn()
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS coconut_count_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                time TEXT NOT NULL,
                total INTEGER NOT NULL,
                belt1 INTEGER NOT NULL,
                belt2 INTEGER NOT NULL,
                belt3 INTEGER NOT NULL,
                batches INTEGER,
                avg_batch INTEGER,
                conveyor_belt TEXT,
                camera TEXT,
                accuracy TEXT,
                status TEXT NOT NULL,
                year INTEGER NOT NULL,
                month INTEGER NOT NULL,
                day INTEGER NOT NULL,
                hour INTEGER NOT NULL,
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_coconut_history_ymd ON coconut_count_history (year, month, day)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_coconut_history_created ON coconut_count_history (created_at)")
        conn.commit()
    finally:
        conn.close()


def _init_dashboard_efficiency_db():
    """Create dashboard_efficiency table for System Performance Overview chart (operations efficiency %)."""
    conn = _get_db_conn()
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS dashboard_efficiency (
                metric_key TEXT PRIMARY KEY,
                value REAL NOT NULL DEFAULT 0,
                updated_at TEXT
            )
        """)
        conn.commit()
    finally:
        conn.close()


def _init_system_stats_db():
    """Create system_stats table for coconut model start count and other system counters."""
    conn = _get_db_conn()
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS system_stats (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL DEFAULT '0',
                updated_at TEXT
            )
        """)
        conn.commit()
    finally:
        conn.close()


def _get_system_stat(key, default="0"):
    """Return value for key from system_stats table."""
    conn = _get_db_conn()
    try:
        cur = conn.execute("SELECT value FROM system_stats WHERE key = ?", (key,))
        row = cur.fetchone()
        return (row["value"] or default) if row else default
    finally:
        conn.close()


def _inc_system_stat(key):
    """Increment integer value for key; create if missing. Returns new value."""
    conn = _get_db_conn()
    try:
        cur = conn.execute("SELECT value FROM system_stats WHERE key = ?", (key,))
        row = cur.fetchone()
        prev = int(row["value"] or 0) if row else 0
        new_val = prev + 1
        now_iso = datetime.now().isoformat()
        conn.execute(
            """INSERT INTO system_stats (key, value, updated_at) VALUES (?, ?, ?)
               ON CONFLICT(key) DO UPDATE SET value = ?, updated_at = ?""",
            (str(key), str(new_val), now_iso, str(new_val), now_iso),
        )
        conn.commit()
        return new_val
    except Exception as e:
        print(f"[system_stats] inc error: {e}")
        return 0
    finally:
        conn.close()


def _init_system_events_history_db():
    """Create system_events_history table for admin-viewable system event history."""
    conn = _get_db_conn()
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS system_events_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_type TEXT NOT NULL DEFAULT 'general',
                message TEXT NOT NULL,
                level TEXT NOT NULL DEFAULT 'info',
                created_at TEXT NOT NULL,
                year INTEGER NOT NULL,
                month INTEGER NOT NULL,
                day INTEGER NOT NULL
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_system_events_created ON system_events_history (created_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_system_events_ymd ON system_events_history (year, month, day)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_system_events_type ON system_events_history (event_type)")
        conn.commit()
    finally:
        conn.close()


def _init_login_users_db():
    """Create login_users table and seed default admin if no users exist."""
    conn = _get_db_conn()
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS login_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_login_users_username ON login_users (username)")
        cur = conn.execute("SELECT COUNT(*) AS n FROM login_users")
        if cur.fetchone()[0] == 0:
            now = datetime.now().isoformat()
            conn.execute(
                """INSERT INTO login_users (username, password_hash, created_at) VALUES (?, ?, ?)""",
                (DEFAULT_ADMIN_USERNAME, generate_password_hash(DEFAULT_ADMIN_PASSWORD), now),
            )
            conn.commit()
    finally:
        conn.close()


def _get_user_for_login(username):
    """Return user row (dict-like) if username exists, else None."""
    conn = _get_db_conn()
    try:
        cur = conn.execute("SELECT id, username, password_hash FROM login_users WHERE username = ?", (username,))
        row = cur.fetchone()
        return dict(row) if row else None
    finally:
        conn.close()


def _init_truck_verification_db():
    """Create truck_verification_images table for storing driver/plate/weight snap images (BLOB)."""
    conn = _get_db_conn()
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS truck_verification_images (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                capture_type TEXT NOT NULL,
                image_data BLOB NOT NULL,
                plate_text TEXT,
                camera_id INTEGER,
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_truck_verif_type ON truck_verification_images (capture_type)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_truck_verif_created ON truck_verification_images (created_at)")
        conn.commit()
    finally:
        conn.close()


TRUCK_VERIFICATION_IMAGES_MAX = 2000


def _get_truck_verification_summary_from_db():
    """Return counts for truck summary page: total_plates_processed, total_driver_snaps, total_trucks, total_weight_snaps."""
    conn = _get_db_conn()
    try:
        cur = conn.execute(
            """SELECT capture_type, COUNT(*) AS n FROM truck_verification_images
               WHERE capture_type IN ('plate_front', 'plate_back', 'driver_snap', 'weight_snap') GROUP BY capture_type"""
        )
        counts = {row["capture_type"]: row["n"] for row in cur.fetchall()}
        plates = int(counts.get("plate_front", 0)) + int(counts.get("plate_back", 0))
        drivers = int(counts.get("driver_snap", 0))
        weights = int(counts.get("weight_snap", 0))
        return {
            "total_plates_processed": plates,
            "total_driver_snaps": drivers,
            "total_trucks": drivers,
            "total_weight_snaps": weights,
            "total_weight": weights,  # alias for truck_summary
        }
    finally:
        conn.close()


def _get_truck_verification_records_from_db(limit=100, from_date=None, to_date=None):
    """Return list of truck records for summary table: each from a driver_snap with matched plate_front, plate_back, weight_snap by time.
    from_date, to_date: optional YYYY-MM-DD strings to filter by date(created_at)."""
    conn = _get_db_conn()
    try:
        query = """SELECT id, created_at FROM truck_verification_images
                    WHERE capture_type = 'driver_snap'"""
        params = []
        if from_date:
            query += " AND date(created_at) >= date(?)"
            params.append(from_date)
        if to_date:
            query += " AND date(created_at) <= date(?)"
            params.append(to_date)
        query += " ORDER BY created_at DESC LIMIT ?"
        params.append(min(limit, 2000))
        cur = conn.execute(query, params)
        driver_rows = cur.fetchall()
        records = []
        for dr in driver_rows:
            d_id = dr["id"]
            d_ts = dr["created_at"]
            # Latest plate_front with created_at <= driver time
            cur2 = conn.execute(
                """SELECT id, plate_text FROM truck_verification_images
                   WHERE capture_type = 'plate_front' AND created_at <= ? ORDER BY created_at DESC LIMIT 1""",
                (d_ts,),
            )
            pf = cur2.fetchone()
            cur2 = conn.execute(
                """SELECT id, plate_text FROM truck_verification_images
                   WHERE capture_type = 'plate_back' AND created_at <= ? ORDER BY created_at DESC LIMIT 1""",
                (d_ts,),
            )
            pb = cur2.fetchone()
            cur2 = conn.execute(
                """SELECT id FROM truck_verification_images
                   WHERE capture_type = 'weight_snap' AND created_at <= ? ORDER BY created_at DESC LIMIT 1""",
                (d_ts,),
            )
            ww = cur2.fetchone()
            plate_front_id = pf["id"] if pf else None
            plate_front_text = (pf["plate_text"] or "").strip() if pf else ""
            plate_back_id = pb["id"] if pb else None
            plate_back_text = (pb["plate_text"] or "").strip() if pb else ""
            weight_snap_id = ww["id"] if ww else None
            records.append({
                "driver_snap_id": d_id,
                "created_at": d_ts,
                "plate_front_id": plate_front_id,
                "plate_front_text": plate_front_text or None,
                "plate_back_id": plate_back_id,
                "plate_back_text": plate_back_text or None,
                "weight_snap_id": weight_snap_id,
            })
        return records
    finally:
        conn.close()


def _get_truck_verification_image_by_id(img_id):
    """Return (image_data, mimetype) for truck_verification_images.id or (None, None)."""
    conn = _get_db_conn()
    try:
        cur = conn.execute("SELECT image_data FROM truck_verification_images WHERE id = ?", (img_id,))
        row = cur.fetchone()
        if row and row["image_data"]:
            return (row["image_data"], "image/jpeg")
        return (None, None)
    finally:
        conn.close()


def _insert_truck_verification_image(capture_type, image_data, plate_text=None, camera_id=None):
    """Insert one truck verification image into DB. capture_type: driver_snap, plate_front, plate_back, weight_snap."""
    if not image_data:
        return
    conn = _get_db_conn()
    try:
        now = datetime.now().isoformat()
        conn.execute(
            """INSERT INTO truck_verification_images (capture_type, image_data, plate_text, camera_id, created_at)
               VALUES (?, ?, ?, ?, ?)""",
            (str(capture_type), image_data, (plate_text or "").strip() or None, camera_id, now),
        )
        conn.commit()
        # Keep table size bounded: delete oldest if over limit
        cur = conn.execute("SELECT COUNT(*) AS n FROM truck_verification_images")
        n = cur.fetchone()[0]
        if n > TRUCK_VERIFICATION_IMAGES_MAX:
            drop_count = n - TRUCK_VERIFICATION_IMAGES_MAX
            cur = conn.execute(
                "SELECT id FROM truck_verification_images ORDER BY id ASC LIMIT ?",
                (drop_count,),
            )
            ids = [row[0] for row in cur.fetchall()]
            if ids:
                placeholders = ",".join("?" * len(ids))
                conn.execute("DELETE FROM truck_verification_images WHERE id IN (" + placeholders + ")", ids)
                conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"[truck_verification_images] insert failed: {e}")
    finally:
        conn.close()


def _insert_system_event_to_db(message, level="info", event_type="general", created_at=None):
    """Persist one system event to DB for history (thread-safe)."""
    if created_at is None:
        created_at = datetime.now()
    conn = _get_db_conn()
    try:
        conn.execute(
            """INSERT INTO system_events_history (event_type, message, level, created_at, year, month, day)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (str(event_type), str(message), str(level), created_at.isoformat(),
             created_at.year, created_at.month, created_at.day),
        )
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise
    finally:
        conn.close()


def _get_system_events_history_from_db(from_date=None, to_date=None, event_type=None, limit=500):
    """Return list of system events from DB for admin history. Newest first."""
    conn = _get_db_conn()
    try:
        query = """SELECT id, event_type, message, level, created_at, year, month, day
                   FROM system_events_history WHERE 1=1"""
        params = []
        if from_date:
            query += " AND date(created_at) >= date(?)"
            params.append(from_date)
        if to_date:
            query += " AND date(created_at) <= date(?)"
            params.append(to_date)
        if event_type:
            query += " AND event_type = ?"
            params.append(event_type)
        query += " ORDER BY created_at DESC LIMIT ?"
        params.append(min(int(limit), 1000))
        cur = conn.execute(query, params)
        rows = cur.fetchall()
        return [
            {
                "id": r["id"],
                "event_type": r["event_type"],
                "message": r["message"],
                "level": r["level"],
                "time": r["created_at"],
                "year": r["year"],
                "month": r["month"],
                "day": r["day"],
            }
            for r in rows
        ]
    finally:
        conn.close()


def _parse_accuracy_to_number(accuracy_str):
    """Parse accuracy string like '99.2%' or '98' to float; return 0 if invalid."""
    if not accuracy_str or accuracy_str == "—":
        return 0.0
    s = str(accuracy_str).strip().replace("%", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def _get_operations_efficiency_from_db():
    """
    Return operations efficiency percentages for dashboard chart from database.
    Keys: coconut_count (from coconut_count_history today's latest accuracy),
          driver_capture, weight_check, license_plate (from dashboard_efficiency table).
    """
    now = datetime.now()
    result = {"coconut_count": 0.0, "driver_capture": 0.0, "weight_check": 0.0, "license_plate": 0.0}
    conn = _get_db_conn()
    try:
        cur = conn.execute(
            """SELECT accuracy FROM coconut_count_history WHERE year = ? AND month = ? AND day = ?
               ORDER BY created_at DESC LIMIT 1""",
            (now.year, now.month, now.day),
        )
        row = cur.fetchone()
        if row and row["accuracy"]:
            result["coconut_count"] = _parse_accuracy_to_number(row["accuracy"])
        cur = conn.execute("SELECT metric_key, value FROM dashboard_efficiency")
        for r in cur.fetchall():
            key = (r["metric_key"] or "").strip()
            if key in result:
                try:
                    result[key] = float(r["value"] or 0)
                except (TypeError, ValueError):
                    pass
    finally:
        conn.close()
    return result


def _upsert_dashboard_efficiency(metric_key, value):
    """Insert or update one efficiency metric for the dashboard chart."""
    conn = _get_db_conn()
    try:
        now_iso = datetime.now().isoformat()
        conn.execute(
            """INSERT INTO dashboard_efficiency (metric_key, value, updated_at) VALUES (?, ?, ?)
               ON CONFLICT(metric_key) DO UPDATE SET value = ?, updated_at = ?""",
            (metric_key, float(value), now_iso, float(value), now_iso),
        )
        conn.commit()
    except Exception as e:
        print(f"[dashboard_efficiency] upsert error: {e}")
    finally:
        conn.close()


def _insert_coconut_snapshot(row):
    """Insert one coconut count snapshot into the database. Thread-safe."""
    conn = _get_db_conn()
    try:
        now = datetime.now()
        created_at = now.isoformat()
        conn.execute("""
            INSERT INTO coconut_count_history
            (date, time, total, belt1, belt2, belt3, batches, avg_batch, conveyor_belt, camera, accuracy, status, year, month, day, hour, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            row["date"],
            row["time"],
            int(row["total"]),
            int(row["belt1"]),
            int(row["belt2"]),
            int(row["belt3"]),
            int(row.get("batches", 0)),
            int(row.get("avgBatch", 0)),
            str(row.get("conveyorBelt", "")),
            str(row.get("camera", "")),
            str(row.get("accuracy", "")),
            str(row["status"]),
            int(row["year"]),
            int(row["month"]),
            int(row["day"]),
            int(row.get("hour", now.hour)),
            created_at,
        ))
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"[coconut_db] insert error: {e}")
    finally:
        conn.close()


def _get_coconut_history_from_db(year=None, month=None, day=None, time_filter=None):
    """
    Return list of dicts for API/PDF. time_filter: None, 'shift1', 'shift2', 'shift3'.
    Order: oldest first (created_at asc) so front-end can show newest last.
    """
    conn = _get_db_conn()
    try:
        query = "SELECT date, time, total, belt1, belt2, belt3, batches, avg_batch, conveyor_belt, camera, accuracy, status, year, month, day, hour FROM coconut_count_history WHERE 1=1"
        params = []
        if year is not None:
            query += " AND year = ?"
            params.append(int(year))
        if month is not None:
            query += " AND month = ?"
            params.append(int(month))
        if day is not None:
            query += " AND day = ?"
            params.append(int(day))
        query += " ORDER BY created_at ASC"
        cur = conn.execute(query, params)
        rows = cur.fetchall()
        out = []
        for r in rows:
            hour = r["hour"] if r["hour"] is not None else 0
            if time_filter == "shift1" and (hour < 6 or hour >= 14):
                continue
            if time_filter == "shift2" and (hour < 14 or hour >= 22):
                continue
            if time_filter == "shift3" and (hour >= 6 and hour < 22):
                continue
            out.append({
                "date": r["date"],
                "time": r["time"],
                "total": r["total"],
                "belt1": r["belt1"],
                "belt2": r["belt2"],
                "belt3": r["belt3"],
                "batches": r["batches"],
                "avgBatch": r["avg_batch"],
                "conveyorBelt": r["conveyor_belt"] or "",
                "camera": r["camera"] or "",
                "accuracy": r["accuracy"] or "99.2%",
                "status": r["status"],
                "year": r["year"],
                "month": r["month"],
                "day": r["day"],
                "hour": hour,
            })
        return out
    finally:
        conn.close()


def _get_coconut_today_summary_from_db():
    """
    Return today's aggregates from DB: total_today, belt1_today, belt2_today, belt3_today, accuracy (from latest row).
    Used by coconut count page to show metrics from database.
    """
    now = datetime.now()
    conn = _get_db_conn()
    try:
        cur = conn.execute(
            """SELECT SUM(total) AS total_today, SUM(belt1) AS b1, SUM(belt2) AS b2, SUM(belt3) AS b3
               FROM coconut_count_history WHERE year = ? AND month = ? AND day = ?""",
            (now.year, now.month, now.day),
        )
        row = cur.fetchone()
        total_today = int(row["total_today"] or 0)
        belt1_today = int(row["b1"] or 0)
        belt2_today = int(row["b2"] or 0)
        belt3_today = int(row["b3"] or 0)
        cur = conn.execute(
            """SELECT accuracy FROM coconut_count_history WHERE year = ? AND month = ? AND day = ?
               ORDER BY created_at DESC LIMIT 1""",
            (now.year, now.month, now.day),
        )
        last = cur.fetchone()
        accuracy = (last["accuracy"] or "—") if last else "—"
        return {
            "total_today": total_today,
            "belt1_today": belt1_today,
            "belt2_today": belt2_today,
            "belt3_today": belt3_today,
            "accuracy": accuracy,
        }
    finally:
        conn.close()


# Initialize DB on import so table exists under any server (e.g. gunicorn)
try:
    _init_coconut_db()
    _init_dashboard_efficiency_db()
    _init_system_stats_db()
    _init_system_events_history_db()
    _init_login_users_db()
    _init_truck_verification_db()
except Exception as e:
    print(f"[coconut_db] init warning: {e}")

# ---------------- Camera Setup ----------------
# RTSP camera URLs (replace with your actual cameras)
# Note: Camera IDs are 0-10 to match the template
#
# Truck LNPR / Driver Snap page (4 dedicated cameras):
#   6 = License Plate (Front) - front number plate camera
#   7 = License Plate (Back)  - back number plate camera
#   8 = Driver Snap           - driver face capture camera
#   9 = Weight Machine        - weight scale / weighbridge camera

camera_urls = {
    0: "rtsp://admin:03121135508dD@192.168.1.64:554/Streaming/Channels/102",  # default / coconut
    1: "rtsp://admin:03121135508dD@192.168.1.64:554/Streaming/Channels/102",
    2: "rtsp://admin:03121135508dD@192.168.1.64:554/Streaming/Channels/102",
    3: "rtsp://admin:03121135508dD@192.168.1.64:554/Streaming/Channels/102",
    4: "rtsp://admin:03121135508dD@192.168.1.64:554/Streaming/Channels/102",
    5: "rtsp://admin:03121135508dD@192.168.1.64:554/Streaming/Channels/102",
    6: "rtsp://admin:03121135508dD@192.168.1.64:554/Streaming/Channels/102",   # License Plate Front
    7: "rtsp://admin:03121135508dD@192.168.1.64:554/Streaming/Channels/102",   # License Plate Back
    8: "rtsp://admin:03121135508dD@192.168.1.64:554/Streaming/Channels/102",   # Driver Snap
    9: "rtsp://admin:03121135508dD@192.168.1.64:554/Streaming/Channels/102",   # Weight Machine
    10: "rtsp://admin:03121135508dD@192.168.1.64:554/Streaming/Channels/102",
}


# ---------------- Unified camera handling: one thread per unique RTSP URL ----------------
# When multiple camera IDs share the same URL (e.g. 6,7,8,9), one connection feeds all to avoid
# "camera not open" from devices that limit concurrent RTSP connections. Use get_frame(camera_id).

def _build_url_to_cams():
    """Map each RTSP URL to list of camera IDs that use it (so one thread per URL)."""
    m = {}
    for cid, url in camera_urls.items():
        m.setdefault(url, []).append(cid)
    return m


_url_to_cam_ids = _build_url_to_cams()
_camera_frame_buffers = {}   # cam_id -> (frame_bgr, timestamp) or (None, 0)
_camera_buffer_locks = {i: threading.Lock() for i in range(11)}
_url_threads = {}             # url -> thread (one reader per URL)
_url_threads_lock = threading.Lock()
RECONNECT_DELAY_SEC = 2


def _camera_reader_thread_for_url(url, cam_ids):
    """Single thread per URL: read RTSP once and write latest frame to all cam_ids that use this URL."""
    while True:
        cap = None
        try:
            cap = cv2.VideoCapture(url, cv2.CAP_FFMPEG)
            cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
            cap.set(cv2.CAP_PROP_OPEN_TIMEOUT_MSEC, 5000)
            cap.set(cv2.CAP_PROP_READ_TIMEOUT_MSEC, 5000)
            if not cap.isOpened():
                raise RuntimeError("open failed")
            for _ in range(8):
                cap.grab()
        except Exception as e:
            if cap is not None:
                try:
                    cap.release()
                except Exception:
                    pass
            for cid in cam_ids:
                with _camera_buffer_locks[cid]:
                    _camera_frame_buffers[cid] = (None, 0)
            print(f"Camera(s) {cam_ids} connect failed: {e}; reconnecting in {RECONNECT_DELAY_SEC}s")
            time.sleep(RECONNECT_DELAY_SEC)
            continue
        print(f"Camera(s) {cam_ids} connected (1 stream)")
        ts = time.time()
        while True:
            ret, frame = cap.read()
            if not ret or frame is None:
                break
            ts = time.time()
            f = frame.copy()
            for cid in cam_ids:
                with _camera_buffer_locks[cid]:
                    _camera_frame_buffers[cid] = (f, ts)
            time.sleep(0)
        try:
            cap.release()
        except Exception:
            pass
        for cid in cam_ids:
            with _camera_buffer_locks[cid]:
                _camera_frame_buffers[cid] = (None, 0)
        print(f"Camera(s) {cam_ids} feed lost; reconnecting in {RECONNECT_DELAY_SEC}s")
        time.sleep(RECONNECT_DELAY_SEC)


def _ensure_camera_thread(cam_id):
    """Start the background reader for this camera's URL if not already running (one thread per URL)."""
    if cam_id not in camera_urls:
        return
    url = camera_urls[cam_id]
    cam_ids = _url_to_cam_ids.get(url, [cam_id])
    with _url_threads_lock:
        if url in _url_threads and _url_threads[url].is_alive():
            return
        for cid in cam_ids:
            _camera_frame_buffers[cid] = (None, 0)
        t = threading.Thread(target=_camera_reader_thread_for_url, args=(url, cam_ids), daemon=True)
        t.start()
        _url_threads[url] = t


def get_frame(camera_id):
    """Return the latest frame for the given camera (thread-safe), or None if unavailable."""
    if camera_id not in camera_urls:
        return None
    _ensure_camera_thread(camera_id)
    with _camera_buffer_locks[camera_id]:
        data = _camera_frame_buffers.get(camera_id)
    if data is None:
        return None
    frame, _ = data
    if frame is None:
        return None
    return frame.copy()


# Coconut count per camera (for coconut counting page; updated by coconut feed generator)
coconut_counts = {i: 0 for i in range(11)}
coconut_counts_lock = threading.Lock()

# Truck LNPR: detection runs only when user clicks "Start detection" on the page
truck_detection_enabled = False
truck_detection_started_at = None  # when last turned on (for duration on System page)
truck_detection_lock = threading.Lock()

# Truck LNPR: recent number plates from cams 6 (front), 7 (back). Max 30 entries. Each: {"camera_id", "camera_name", "plate", "timestamp"}
recent_plates = []
recent_plates_lock = threading.Lock()
RECENT_PLATES_MAX = 30
TRUCK_CAM_NAMES = {6: "License Plate (Front)", 7: "License Plate (Back)", 8: "Driver Snap", 9: "Weight Machine"}
TRUCK_PLATE_CAMS = (6, 7)
TRUCK_DRIVER_CAM = 8
TRUCK_ALPR_RUN_EVERY_N = 5
# Driver snap: when face/person detected on cam 8, capture crop and store for sidebar
latest_driver_snap = None  # jpeg bytes or None
latest_driver_snap_timestamp = None  # datetime when last captured (for "Captured!" indicator)
latest_driver_snap_lock = threading.Lock()
DRIVER_SNAP_JPEG_QUALITY = 88

# Plate snap: when ALPR detects number plate on cam 6/7, capture plate crop and store for sidebar (front/back separate)
latest_plate_snap = None  # jpeg bytes or None (last updated from either cam)
latest_plate_snap_timestamp = None
latest_plate_snap_text = None
latest_plate_snap_camera_name = None
latest_plate_snap_lock = threading.Lock()
latest_plate_snap_front = None  # cam 6
latest_plate_snap_back = None   # cam 7
# Weight machine: latest frame from cam 9 for right-side grid
latest_weight_frame = None
latest_weight_frame_lock = threading.Lock()
PLATE_SNAP_JPEG_QUALITY = 90
_placeholder_driver_snap = None  # small "No capture yet" jpeg for API when no snap

def _get_placeholder_driver_snap():
    global _placeholder_driver_snap
    if _placeholder_driver_snap is not None:
        return _placeholder_driver_snap
    ph = np.zeros((120, 120, 3), dtype=np.uint8)
    ph[:] = (40, 40, 40)
    cv2.putText(ph, "No driver", (12, 58), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (180, 180, 180), 1)
    cv2.putText(ph, "capture yet", (8, 78), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (180, 180, 180), 1)
    ret, buf = cv2.imencode('.jpg', ph, [cv2.IMWRITE_JPEG_QUALITY, 70])
    _placeholder_driver_snap = buf.tobytes() if ret else b''
    return _placeholder_driver_snap


_placeholder_plate_snap = None


def _get_placeholder_plate_snap():
    global _placeholder_plate_snap
    if _placeholder_plate_snap is not None:
        return _placeholder_plate_snap
    ph = np.zeros((80, 220, 3), dtype=np.uint8)
    ph[:] = (45, 45, 45)
    cv2.putText(ph, "No plate", (50, 32), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (160, 160, 160), 1)
    cv2.putText(ph, "capture yet", (42, 52), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (160, 160, 160), 1)
    ret, buf = cv2.imencode('.jpg', ph, [cv2.IMWRITE_JPEG_QUALITY, 70])
    _placeholder_plate_snap = buf.tobytes() if ret else b''
    return _placeholder_plate_snap


_placeholder_weight_snap = None


def _get_placeholder_weight_snap():
    global _placeholder_weight_snap
    if _placeholder_weight_snap is not None:
        return _placeholder_weight_snap
    ph = np.zeros((120, 160, 3), dtype=np.uint8)
    ph[:] = (45, 45, 45)
    cv2.putText(ph, "Weight", (40, 55), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (160, 160, 160), 1)
    cv2.putText(ph, "camera", (42, 75), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (160, 160, 160), 1)
    ret, buf = cv2.imencode('.jpg', ph, [cv2.IMWRITE_JPEG_QUALITY, 70])
    _placeholder_weight_snap = buf.tobytes() if ret else b''
    return _placeholder_weight_snap


# Coconut detection on/off: when False, 6 cameras show raw feed and count 0; when True, run model and show boxes + count
coconut_detection_enabled = False
coconut_detection_lock = threading.Lock()
# When detection was last turned on (datetime), for duration display and PDF report
coconut_detection_started_at = None

# Count history: rows appended every 30s while detection on, and when detection turns off (for table + PDF)
COCONUT_HISTORY_MAX = 500
coconut_count_history = []
coconut_count_history_lock = threading.Lock()
_last_history_append_at = None

def _append_coconut_count_snapshot():
    """Append current counts to history (call every 30s while detection on, or when detection off)."""
    global _last_history_append_at
    with coconut_counts_lock:
        counts = dict(coconut_counts)
    with coconut_detection_lock:
        enabled = coconut_detection_enabled
    now = datetime.now()
    c0 = counts.get(0, 0) + counts.get(1, 0)
    c1 = counts.get(2, 0) + counts.get(3, 0)
    c2 = counts.get(4, 0) + counts.get(5, 0)
    belt1, belt2, belt3 = c0, c1, c2
    total = belt1 + belt2 + belt3
    batches = max(1, total // 88)
    avg_batch = total // batches
    if total > 0:
        if belt1 >= belt2 and belt1 >= belt3:
            belt_label = "Belt 1"
        elif belt2 >= belt1 and belt2 >= belt3:
            belt_label = "Belt 2"
        else:
            belt_label = "Belt 3"
    else:
        belt_label = "Belt 1-3"
    row = {
        "date": now.strftime("%b %d, %Y"),
        "time": now.strftime("%H:%M"),
        "total": total,
        "belt1": belt1,
        "belt2": belt2,
        "belt3": belt3,
        "batches": batches,
        "avgBatch": avg_batch,
        "conveyorBelt": belt_label,
        "camera": "Cam 1-6",
        "accuracy": "99.2%",
        "status": "Live" if enabled else "Completed",
        "year": now.year,
        "month": now.month,
        "day": now.day,
        "hour": now.hour,
    }
    # Persist to SQLite for weight_check page and reports
    try:
        _insert_coconut_snapshot(row)
        # Update dashboard efficiency chart: coconut_count from latest accuracy
        _upsert_dashboard_efficiency("coconut_count", _parse_accuracy_to_number(row.get("accuracy", "")))
    except Exception as e:
        print(f"[coconut_db] snapshot insert failed: {e}")
    with coconut_count_history_lock:
        coconut_count_history.append(row)
        if len(coconut_count_history) > COCONUT_HISTORY_MAX:
            coconut_count_history.pop(0)
    _last_history_append_at = now


def _coconut_history_loop():
    """Every 30s if detection on, append a snapshot to history."""
    while True:
        time.sleep(30)
        with coconut_detection_lock:
            if not coconut_detection_enabled:
                continue
        _append_coconut_count_snapshot()


# ---------------- System cache clear every 6 hours (for performance) ----------------
CACHE_CLEAR_INTERVAL_SEC = 6 * 3600  # 6 hours


def clear_all_caches():
    """Remove in-memory caches: recent plates, coconut detection cache, coconut count history, reset coconut counts. Then run gc."""
    with recent_plates_lock:
        recent_plates.clear()
    with _last_detection_result_lock:
        _last_detection_result.clear()
    with coconut_count_history_lock:
        coconut_count_history.clear()
    with coconut_counts_lock:
        for k in coconut_counts:
            coconut_counts[k] = 0
    gc.collect()


def _cache_clear_loop():
    """Every 6 hours, clear all caches for system performance."""
    while True:
        time.sleep(CACHE_CLEAR_INTERVAL_SEC)
        try:
            clear_all_caches()
            _log_system_event("Cache cleared automatically (every 6 hours) for performance.", "info", event_type="cache_clear")
        except Exception as e:
            _log_system_event("Auto cache clear failed: " + str(e), "error", event_type="cache_clear")


# Single inference lock: only one YOLO detection at a time across all 6 cameras (smooth, no GPU contention)
coconut_inference_lock = threading.Lock()
# Per-camera cache of last detection result so we can show smooth video while running detection every Nth frame
_last_detection_result = {}  # cam_id -> (annotated_frame, count)
_last_detection_result_lock = threading.Lock()
DETECTION_RUN_EVERY_N_FRAMES = 2  # run model every 2nd frame per camera, show cached result otherwise

# ---------------- Helper Functions ----------------
def generate_frames(cam_id):
    """Generate MJPEG frames from the latest camera frame (uses get_frame from background thread)."""
    while True:
        frame = get_frame(cam_id)
        if frame is None:
            frame = create_no_signal_frame()
            time.sleep(1)
        ret, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 85])
        if ret:
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
        time.sleep(0.04)

def create_no_signal_frame():
    """Create a placeholder frame when camera is not available (use numpy; cv2 has no zeros)."""
    frame = np.zeros((480, 640, 3), dtype=np.uint8)
    cv2.putText(frame, "NO SIGNAL", (180, 220),
                cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
    cv2.putText(frame, "Camera Offline", (200, 260),
                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
    return frame


# ---------------- Truck LNPR / Driver Snap feeds (ALPR on 6/7, person on 8, raw 9) ----------------
def generate_truck_frames(cam_id):
    """Generate MJPEG: cams 6/7 = ALPR + EasyOCR + recent_plates + plate capture for sidebar; cam 8 = face/person + capture; cam 9 = raw. Detection runs only when truck_detection_enabled is True."""
    global latest_driver_snap, latest_driver_snap_timestamp
    global latest_plate_snap, latest_plate_snap_timestamp, latest_plate_snap_text, latest_plate_snap_camera_name
    global latest_plate_snap_front, latest_plate_snap_back, latest_weight_frame
    if cam_id not in (6, 7, 8, 9):
        while True:
            frame = create_no_signal_frame()
            ret, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 85])
            if ret:
                yield (b'--frame\r\n' b'Content-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
            time.sleep(0.04)
        return
    try:
        import truck_alpr
    except Exception:
        truck_alpr = None
    frame_count = 0
    jpeg_quality = 85
    while True:
        frame = get_frame(cam_id)
        if frame is None:
            frame = create_no_signal_frame()
            ret, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, jpeg_quality])
            if ret:
                yield (b'--frame\r\n' b'Content-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
            time.sleep(0.04)
            continue
        with truck_detection_lock:
            detection_on = truck_detection_enabled
        if detection_on and cam_id in TRUCK_PLATE_CAMS and truck_alpr is not None:
            frame_count += 1
            if frame_count % TRUCK_ALPR_RUN_EVERY_N == 0:
                try:
                    annotated, plate_text, plate_crop = truck_alpr.detect_and_ocr(frame, conf_threshold=0.35)
                    frame = annotated
                    name = TRUCK_CAM_NAMES.get(cam_id, f"Cam {cam_id}")
                    # When number plate is detected: show crop on right panel and add to recent list
                    if plate_crop is not None:
                        ret_p, buf_p = cv2.imencode('.jpg', plate_crop, [cv2.IMWRITE_JPEG_QUALITY, PLATE_SNAP_JPEG_QUALITY])
                        if ret_p:
                            buf_bytes = buf_p.tobytes()
                            with latest_plate_snap_lock:
                                latest_plate_snap = buf_bytes
                                latest_plate_snap_timestamp = datetime.now()
                                latest_plate_snap_text = plate_text if plate_text else ""
                                latest_plate_snap_camera_name = name
                                if cam_id == 6:
                                    latest_plate_snap_front = buf_bytes
                                else:
                                    latest_plate_snap_back = buf_bytes
                            try:
                                _insert_truck_verification_image(
                                    'plate_front' if cam_id == 6 else 'plate_back',
                                    buf_bytes,
                                    plate_text=plate_text if plate_text else "",
                                    camera_id=cam_id,
                                )
                            except Exception:
                                pass
                        with recent_plates_lock:
                            recent_plates.append({
                                "camera_id": cam_id,
                                "camera_name": name,
                                "plate": plate_text if plate_text else "—",
                                "timestamp": datetime.now().isoformat(),
                            })
                            if len(recent_plates) > RECENT_PLATES_MAX:
                                recent_plates.pop(0)
                except Exception:
                    pass
        elif detection_on and cam_id == TRUCK_DRIVER_CAM and truck_alpr is not None:
            try:
                frame, driver_crop = truck_alpr.detect_person_and_crop(frame, conf_threshold=0.35)
                if driver_crop is not None:
                    ret_j, buf_j = cv2.imencode('.jpg', driver_crop, [cv2.IMWRITE_JPEG_QUALITY, DRIVER_SNAP_JPEG_QUALITY])
                    if ret_j:
                        driver_bytes = buf_j.tobytes()
                        with latest_driver_snap_lock:
                            latest_driver_snap = driver_bytes
                            latest_driver_snap_timestamp = datetime.now()
                        try:
                            _insert_truck_verification_image('driver_snap', driver_bytes)
                            with latest_weight_frame_lock:
                                wf = latest_weight_frame
                            if wf:
                                _insert_truck_verification_image('weight_snap', wf)
                        except Exception:
                            pass
            except Exception:
                pass
        ret, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, jpeg_quality])
        if ret:
            if cam_id == 9:
                with latest_weight_frame_lock:
                    latest_weight_frame = buffer.tobytes()
            yield (b'--frame\r\n' b'Content-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
        time.sleep(0.04)


# ---------------- Coconut detection feed (model from train_models/ on 6 cameras) ----------------
def generate_coconut_frames(cam_id):
    """Generate frames from get_frame(cam_id); detection when enabled with single inference lock + cache for smooth 6-cam."""
    from coconut_count_model import detect_coconuts

    last_ts = 0
    jpeg_quality = 82
    frame_count = 0

    while True:
        frame = get_frame(cam_id)
        if frame is None:
            with coconut_counts_lock:
                coconut_counts[cam_id] = 0
            frame = create_no_signal_frame()
        # else: frame is already the latest from get_frame(cam_id)

        with coconut_detection_lock:
            detection_on = coconut_detection_enabled

        if detection_on:
            frame_count += 1
            run_detection_this_frame = (frame_count % DETECTION_RUN_EVERY_N_FRAMES == 1)
            if run_detection_this_frame:
                with coconut_inference_lock:
                    annotated, count = detect_coconuts(frame, conf_threshold=0.25, iou_threshold=0.45)
                cv2.putText(
                    annotated, f"Coconuts: {count}",
                    (10, 40), cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 255, 100), 2, cv2.LINE_AA
                )
                with _last_detection_result_lock:
                    _last_detection_result[cam_id] = (annotated.copy(), count)
                with coconut_counts_lock:
                    coconut_counts[cam_id] = count
                out_frame = annotated
            else:
                with _last_detection_result_lock:
                    cached = _last_detection_result.get(cam_id)
                if cached is not None:
                    out_frame, count = cached
                    with coconut_counts_lock:
                        coconut_counts[cam_id] = count
                    out_frame = out_frame.copy()
                else:
                    with coconut_inference_lock:
                        annotated, count = detect_coconuts(frame, conf_threshold=0.25, iou_threshold=0.45)
                    cv2.putText(
                        annotated, f"Coconuts: {count}",
                        (10, 40), cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 255, 100), 2, cv2.LINE_AA
                    )
                    with _last_detection_result_lock:
                        _last_detection_result[cam_id] = (annotated.copy(), count)
                    with coconut_counts_lock:
                        coconut_counts[cam_id] = count
                    out_frame = annotated
        else:
            with coconut_counts_lock:
                coconut_counts[cam_id] = 0
            with _last_detection_result_lock:
                _last_detection_result.pop(cam_id, None)
            out_frame = frame

        ret, buffer = cv2.imencode('.jpg', out_frame, [cv2.IMWRITE_JPEG_QUALITY, jpeg_quality])
        if ret:
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
        t = time.time()
        # Throttle: when detection on, moderate interval so lock shared fairly (faster model = smoother)
        throttle = 0.045 if detection_on else 0.04
        elapsed = t - last_ts
        if elapsed < throttle:
            time.sleep(throttle - elapsed)
        last_ts = time.time()

# ---------------- Login / Logout ----------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        user = _get_user_for_login(username)
        if user and check_password_hash(user["password_hash"], password):
            session["logged_in"] = True
            session["username"] = user["username"]
            return redirect(url_for("home"))
        return render_template("login.html", error="Invalid username or password.")
    if session.get("logged_in"):
        return redirect(url_for("home"))
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.pop("logged_in", None)
    session.pop("username", None)
    return redirect(url_for("login"))


@app.before_request
def require_login():
    """Always show login first: redirect to login if not logged in (except login, logout, static)."""
    if request.path.startswith("/static/"):
        return
    if request.path in ("/login", "/logout"):
        return
    if not session.get("logged_in"):
        return redirect(url_for("login"))


# ---------------- Page Routes ----------------
@app.route("/")
def home():
    return render_template("index.html")  # Dashboard

@app.route("/live_camera")
def live_camera():
    return render_template("live_camera.html")  # 10-camera live view

@app.route("/coconut_count")
def coconut_count():
    return render_template("coconut_count.html")  # Coconut counting analytics dashboard

@app.route("/truck_lnpr_driver_snap")
def truck_lnpr_driver_snap():
    return render_template("truck_lnpr_driver_snap.html")  # Truck verification & security compliance

@app.route("/weight_check")
def weight_check():
    return render_template("weight_check.html")  # Coconut Summary

@app.route("/truck_summary")
def truck_summary():
    return render_template("truck_summary.html")  # Truck Summary with filters


@app.route("/system")
def system_page():
    return render_template("system.html")  # System & project details, event log

# ---------------- Download Routes ----------------
@app.route("/download_truck_data_excel", methods=["POST"])
def download_truck_data_excel():
    if not OPENPYXL_AVAILABLE:
        return jsonify({"error": "openpyxl library not installed. Please install it using: pip install openpyxl"}), 500
    
    try:
        data = request.get_json()
        truck_data = data.get('truckData', [])
        report_date = data.get('reportDate', datetime.now().strftime('%Y-%m-%d'))
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Truck Summary"
        
        # Define styles
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"Truck Summary Report - {report_date}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Add headers
        headers = ['#', 'Number Plate', 'Plate Image URL', 'Driver Image URL', 'Weight (kg)', 'Date', 'Time', 'Status']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Add data
        for row_num, truck in enumerate(truck_data, 4):
            ws.cell(row=row_num, column=1, value=row_num - 3).border = border
            ws.cell(row=row_num, column=2, value=truck.get('plateNumber', '')).border = border
            ws.cell(row=row_num, column=3, value=truck.get('plateImage', '')).border = border
            ws.cell(row=row_num, column=4, value=truck.get('driverImage', '')).border = border
            ws.cell(row=row_num, column=5, value=truck.get('weight', 0)).border = border
            ws.cell(row=row_num, column=6, value=truck.get('date', '')).border = border
            ws.cell(row=row_num, column=7, value=truck.get('time', '')).border = border
            ws.cell(row=row_num, column=8, value=truck.get('status', '')).border = border
        
        # Adjust column widths
        column_widths = [5, 18, 35, 35, 12, 12, 12, 12]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Create in-memory file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"truck_data_{report_date}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/download_truck_data_pdf", methods=["POST"])
def download_truck_data_pdf():
    if not REPORTLAB_AVAILABLE:
        return jsonify({"error": "reportlab library not installed. Please install it using: pip install reportlab"}), 500
    
    try:
        data = request.get_json()
        truck_data = data.get('truckData', [])
        report_date = data.get('reportDate', datetime.now().strftime('%Y-%m-%d'))
        
        # Create in-memory PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e40af'),
            alignment=1,  # Center alignment
            spaceAfter=30
        )
        
        # Add title
        title = Paragraph(f"Truck Summary Report - {report_date}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Prepare table data
        table_data = [['#', 'Number Plate', 'Weight (kg)', 'Date', 'Time', 'Status']]
        
        for idx, truck in enumerate(truck_data, 1):
            table_data.append([
                str(idx),
                truck.get('plateNumber', ''),
                str(truck.get('weight', 0)),
                truck.get('date', ''),
                truck.get('time', ''),
                truck.get('status', '').upper()
            ])
        
        # Create table
        table = Table(table_data)
        
        # Add style to table
        table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            # Data rows
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        
        # Add summary
        elements.append(Spacer(1, 0.3*inch))
        total_weight = sum(truck.get('weight', 0) for truck in truck_data)
        summary_text = f"<b>Total Trucks:</b> {len(truck_data)} | <b>Total Weight:</b> {total_weight:,} kg"
        summary = Paragraph(summary_text, styles['Normal'])
        elements.append(summary)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"truck_report_{report_date}.pdf"
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def _truck_report_period_from_request():
    """Parse period=daily|monthly|days and date/month/year/days. Returns (from_date, to_date, label). No yearly."""
    period = (request.args.get("period") or "daily").strip().lower()
    today = datetime.now().date()
    from_date = to_date = None
    label = "Report"
    if period == "daily":
        date_s = request.args.get("date")
        if date_s:
            try:
                from datetime import date as date_cls
                d = date_cls.fromisoformat(date_s)
                from_date = to_date = d.isoformat()
                label = "Daily " + from_date
            except (ValueError, TypeError):
                pass
        if not from_date:
            from_date = to_date = today.isoformat()
            label = "Daily " + from_date
    elif period == "monthly":
        try:
            y = int(request.args.get("year") or today.year)
            m = int(request.args.get("month") or today.month)
        except (TypeError, ValueError):
            y, m = today.year, today.month
        m = max(1, min(12, m))
        from_date = "%04d-%02d-01" % (y, m)
        if m == 12:
            to_date = "%04d-12-31" % y
        else:
            from datetime import date as date_cls, timedelta
            to_date = (date_cls(y, m + 1, 1) - timedelta(days=1)).isoformat()
        month_names = ("", "January", "February", "March", "April", "May", "June",
                      "July", "August", "September", "October", "November", "December")
        label = "%s %s" % (month_names[m], y)
    elif period == "days":
        try:
            n = int(request.args.get("days") or 7)
            n = max(1, min(365, n))
        except (TypeError, ValueError):
            n = 7
        from datetime import timedelta
        to_date = today.isoformat()
        from_date = (today - timedelta(days=n - 1)).isoformat()
        label = "Last %d days" % n
    else:
        from_date = to_date = today.isoformat()
        label = "Daily " + from_date
    return from_date, to_date, label


@app.route("/download_truck_report_pdf")
def download_truck_report_pdf():
    """Download truck summary PDF by period: period=daily&date=YYYY-MM-DD, or period=monthly&month=M&year=Y, or period=days&days=N. No yearly."""
    if not REPORTLAB_AVAILABLE:
        return jsonify({"error": "reportlab not installed. pip install reportlab"}), 500
    from_date, to_date, label = _truck_report_period_from_request()
    try:
        records = _get_truck_verification_records_from_db(limit=2000, from_date=from_date, to_date=to_date)
    except Exception as e:
        return jsonify({"error": "Failed to load truck data: " + str(e)}), 500

    def _pdf_image_or_placeholder(img_id, w, h):
        """Return a ReportLab Image flowable from DB blob, or Paragraph 'No image' if missing."""
        if not img_id:
            return Paragraph("<font size=6>No image</font>", ParagraphStyle("Tiny", fontName="Helvetica", fontSize=6, alignment=1))
        data, _ = _get_truck_verification_image_by_id(img_id)
        if data:
            try:
                return Image(io.BytesIO(data), width=w, height=h)
            except Exception:
                return Paragraph("<font size=6>No image</font>", ParagraphStyle("Tiny", fontName="Helvetica", fontSize=6, alignment=1))
        return Paragraph("<font size=6>No image</font>", ParagraphStyle("Tiny", fontName="Helvetica", fontSize=6, alignment=1))

    img_w_plate = 1.0 * inch
    img_h_plate = 0.5 * inch
    img_w_driver = 0.65 * inch
    img_h_driver = 0.65 * inch

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=20,
        textColor=colors.HexColor("#1e40af"),
        alignment=1,
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "SubTitle",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#6b7280"),
        alignment=1,
        spaceAfter=20,
    )
    elements.append(Paragraph("Truck Summary Report", title_style))
    elements.append(Paragraph("%s &middot; Generated %s" % (label, datetime.now().strftime("%Y-%m-%d %H:%M")), subtitle_style))
    table_data = [["#", "Number Plate", "Plate Image", "Driver Image", "Weight (kg)", "Date", "Time", "Status"]]
    for idx, r in enumerate(records, 1):
        plate_number = (r.get("plate_front_text") or r.get("plate_back_text") or "").strip() or "—"
        created_at = r.get("created_at")
        if created_at:
            try:
                dt = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                if dt.tzinfo:
                    dt = dt.replace(tzinfo=None)
                date_str = dt.strftime("%Y-%m-%d")
                time_str = dt.strftime("%H:%M:%S")
            except Exception:
                date_str = created_at[:10] if len(created_at) >= 10 else "—"
                time_str = created_at[11:19] if len(created_at) >= 19 else "—"
        else:
            date_str = time_str = "—"
        plate_id = r.get("plate_front_id") or r.get("plate_back_id")
        driver_id = r.get("driver_snap_id")
        table_data.append([
            str(idx),
            plate_number,
            _pdf_image_or_placeholder(plate_id, img_w_plate, img_h_plate),
            _pdf_image_or_placeholder(driver_id, img_w_driver, img_h_driver),
            "0",
            date_str,
            time_str,
            "COMPLETED",
        ])
    col_widths = [0.35*inch, 1.15*inch, img_w_plate + 0.1*inch, img_w_driver + 0.1*inch, 0.6*inch, 0.85*inch, 0.65*inch, 0.7*inch]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
        ("TOPPADDING", (0, 0), (-1, 0), 10),
        ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#374151")),
        ("LINEBELOW", (0, 0), (-1, 0), 1.5, colors.HexColor("#1e3a8a")),
        ("LINEABOVE", (0, 1), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#fafafa")),
        ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#1f2937")),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (1, -1), 9),
        ("FONTSIZE", (4, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (2, 0), (3, -1), 6),
        ("BOTTOMPADDING", (2, 0), (3, -1), 6),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.35*inch))
    summary_style = ParagraphStyle(
        "Summary",
        parent=styles["Normal"],
        fontSize=11,
        textColor=colors.HexColor("#374151"),
        leftIndent=0,
        spaceAfter=0,
    )
    summary_text = "<b>Total Trucks:</b> %s" % len(records)
    elements.append(Paragraph(summary_text, summary_style))
    doc.build(elements)
    buffer.seek(0)
    safe_label = "".join(c if c.isalnum() or c in "-_" else "_" for c in label)
    filename = "truck_report_%s_%s.pdf" % (safe_label.replace(" ", "_"), datetime.now().strftime("%Y-%m-%d_%H%M"))
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=filename)


# ---------------- Video Feed Routes ----------------
@app.route("/video_feed/<int:cam_id>")
def video_feed(cam_id):
    if cam_id not in camera_urls:
        return "Camera not found", 404
    return Response(generate_frames(cam_id),
                    mimetype='multipart/x-mixed-replace; boundary=frame')


@app.route("/coconut_video_feed/<int:cam_id>")
def coconut_video_feed(cam_id):
    """Coconut counting cameras: YOLO coconut.pt detection with bounding boxes."""
    if cam_id not in camera_urls:
        return "Camera not found", 404
    return Response(generate_coconut_frames(cam_id),
                    mimetype='multipart/x-mixed-replace; boundary=frame')


@app.route("/api/coconut_counts")
def api_coconut_counts():
    """Return current coconut count per camera for the coconut count page."""
    with coconut_counts_lock:
        return jsonify({str(k): v for k, v in coconut_counts.items()})


@app.route("/truck_video_feed/<int:cam_id>")
def truck_video_feed(cam_id):
    """Truck LNPR page: ALPR+OCR on 6/7, person detection on 8, raw on 9."""
    if cam_id not in (6, 7, 8, 9):
        return "Camera not found", 404
    return Response(generate_truck_frames(cam_id),
                    mimetype='multipart/x-mixed-replace; boundary=frame')


@app.route("/api/truck/detection/start", methods=["POST"])
def api_truck_detection_start():
    """Enable truck LNPR/driver detection (ALPR on 6/7, person on 8). Model runs only when enabled."""
    global truck_detection_enabled, truck_detection_started_at
    with truck_detection_lock:
        truck_detection_enabled = True
        truck_detection_started_at = datetime.now()
    start_count = _inc_system_stat("truck_model_start_count")
    _log_system_event("Truck (LNPR/Driver) detection model started (start #%d)." % start_count, "success", event_type="truck_model_start")
    return jsonify({"ok": True, "enabled": True})


@app.route("/api/truck/detection/stop", methods=["POST"])
def api_truck_detection_stop():
    """Disable truck LNPR/driver detection; feeds continue with raw video only."""
    global truck_detection_enabled, truck_detection_started_at
    with truck_detection_lock:
        truck_detection_enabled = False
        truck_detection_started_at = None
    stop_count = _inc_system_stat("truck_model_stop_count")
    _log_system_event("Truck (LNPR/Driver) detection model stopped (stop #%d)." % stop_count, "info", event_type="truck_model_stop")
    return jsonify({"ok": True, "enabled": False})


@app.route("/api/truck/detection/status")
def api_truck_detection_status():
    """Return whether truck detection is currently enabled."""
    with truck_detection_lock:
        enabled = truck_detection_enabled
    return jsonify({"enabled": enabled})


@app.route("/api/truck/recent_plates")
def api_truck_recent_plates():
    """Return recent number plate detections for the truck LNPR sidebar (newest last)."""
    with recent_plates_lock:
        out = list(recent_plates)
    return jsonify(out)


@app.route("/api/truck/driver_snap")
def api_truck_driver_snap():
    """Return latest captured driver image (face/person on cam 8) as JPEG for sidebar. Placeholder if none yet."""
    with latest_driver_snap_lock:
        data = latest_driver_snap
    if data is None:
        data = _get_placeholder_driver_snap()
    return Response(data, mimetype='image/jpeg')


@app.route("/api/truck/driver_snap_info")
def api_truck_driver_snap_info():
    """Return whether a driver was captured and when (for sidebar 'Captured!' indicator)."""
    with latest_driver_snap_lock:
        ts = latest_driver_snap_timestamp
        has = latest_driver_snap is not None
    return jsonify({
        "captured": has,
        "captured_at": ts.isoformat() if ts else None,
    })


@app.route("/api/truck/plate_snap")
def api_truck_plate_snap():
    """Return latest captured number plate image (ALPR on cam 6/7) as JPEG for sidebar. Placeholder if none yet."""
    with latest_plate_snap_lock:
        data = latest_plate_snap
    if data is None:
        data = _get_placeholder_plate_snap()
    return Response(data, mimetype='image/jpeg')


@app.route("/api/truck/plate_snap/<int:cam_id>")
def api_truck_plate_snap_by_cam(cam_id):
    """Return latest plate snap for front (6) or back (7) camera. Placeholder if none yet."""
    if cam_id not in (6, 7):
        return "Not found", 404
    with latest_plate_snap_lock:
        data = latest_plate_snap_front if cam_id == 6 else latest_plate_snap_back
    if data is None:
        data = _get_placeholder_plate_snap()
    return Response(data, mimetype='image/jpeg')


@app.route("/api/truck/weight_snap")
def api_truck_weight_snap():
    """Return latest frame from weight machine camera (9) for right-side grid. Placeholder if none yet."""
    with latest_weight_frame_lock:
        data = latest_weight_frame
    if data is None:
        data = _get_placeholder_weight_snap()
    return Response(data, mimetype='image/jpeg')


@app.route("/api/truck/plate_snap_info")
def api_truck_plate_snap_info():
    """Return latest plate capture info (for sidebar 'Captured' + plate text)."""
    with latest_plate_snap_lock:
        ts = latest_plate_snap_timestamp
        text = latest_plate_snap_text
        name = latest_plate_snap_camera_name
        has = latest_plate_snap is not None
    return jsonify({
        "captured": has,
        "captured_at": ts.isoformat() if ts else None,
        "plate": text or "",
        "camera_name": name or "",
    })


@app.route("/api/truck/verification_summary")
def api_truck_verification_summary():
    """Return summary counts for truck summary page from DB (plates, driver snaps, trucks)."""
    try:
        out = _get_truck_verification_summary_from_db()
        return jsonify(out)
    except Exception as e:
        return jsonify({
            "total_plates_processed": 0,
            "total_driver_snaps": 0,
            "total_trucks": 0,
            "total_weight_snaps": 0,
            "total_weight": 0,
        })


@app.route("/api/truck/verification_records")
def api_truck_verification_records():
    """Return truck verification records from DB for summary table (driver + plate + weight image ids)."""
    limit = request.args.get("limit", 100, type=int)
    limit = min(max(1, limit), 500)
    try:
        records = _get_truck_verification_records_from_db(limit=limit)
        return jsonify({"records": records})
    except Exception as e:
        return jsonify({"records": []})


@app.route("/api/truck/verification_image/<int:img_id>")
def api_truck_verification_image(img_id):
    """Return stored truck verification image by id (JPEG BLOB from DB)."""
    data, mimetype = _get_truck_verification_image_by_id(img_id)
    if data is None:
        return "Image not found", 404
    return Response(data, mimetype=mimetype or "image/jpeg")


@app.route("/api/coconut_detection/status")
def api_coconut_detection_status():
    """Return whether coconut detection is on and when it started (for duration)."""
    with coconut_detection_lock:
        started_at = coconut_detection_started_at
        enabled = coconut_detection_enabled
    out = {"enabled": enabled}
    if started_at is not None:
        out["started_at"] = started_at.isoformat()
    return jsonify(out)


def _preload_coconut_model():
    """Load YOLO model in background so Start detection does not hang on first use."""
    try:
        from coconut_count_model import load_model, get_model_loaded
        if not get_model_loaded():
            load_model()
    except Exception:
        pass


def _preload_truck_alpr_model():
    """Load ALPR model (alpr.pt + EasyOCR) in background so number plate detection is ready on truck page."""
    try:
        import truck_alpr
        truck_alpr.load_alpr_model()
        truck_alpr._init_easyocr()
    except Exception:
        pass


@app.route("/api/coconut_detection/warmup")
def api_coconut_detection_warmup():
    """Trigger model preload in background so detection starts smoothly. Returns immediately."""
    t = threading.Thread(target=_preload_coconut_model, daemon=True)
    t.start()
    return jsonify({"status": "preloading"})


@app.route("/api/coconut_detection/set", methods=["POST"])
def api_coconut_detection_set():
    """Turn coconut detection on or off. Body: {"enabled": true|false}."""
    global coconut_detection_enabled, coconut_detection_started_at
    try:
        data = request.get_json(force=True, silent=True) or {}
        enabled = data.get("enabled", False)
        enabled = bool(enabled)
    except Exception:
        return jsonify({"error": "Invalid JSON"}), 400
    with coconut_detection_lock:
        coconut_detection_enabled = enabled
        if enabled:
            coconut_detection_started_at = datetime.now()
            try:
                start_count = _inc_system_stat("coconut_model_start_count")
                _log_system_event("Coconut detection model started (start #%d)." % start_count, "success", event_type="coconut_model_start")
            except Exception:
                _log_system_event("Coconut detection model started.", "success", event_type="coconut_model_start")
        else:
            coconut_detection_started_at = None
            stop_count = _inc_system_stat("coconut_model_stop_count")
            _log_system_event("Coconut detection model stopped (stop #%d)." % stop_count, "info", event_type="coconut_model_stop")
    if enabled:
        _append_coconut_count_snapshot()
    if not enabled:
        _append_coconut_count_snapshot()
        with _last_detection_result_lock:
            _last_detection_result.clear()
    return jsonify({"enabled": coconut_detection_enabled})


@app.route("/api/coconut_today_summary")
def api_coconut_today_summary():
    """Return today's coconut aggregates from DB for the coconut count page."""
    try:
        summary = _get_coconut_today_summary_from_db()
        return jsonify(summary)
    except Exception as e:
        print(f"[api] coconut_today_summary error: {e}")
        return jsonify({
            "total_today": 0,
            "belt1_today": 0,
            "belt2_today": 0,
            "belt3_today": 0,
            "accuracy": "—",
        })


def _get_production_last_7_days():
    """Return last 7 days (oldest to newest) with daily coconut total and daily truck (driver_snap) count for dashboard chart."""
    today = date.today()
    days = [today - timedelta(days=i) for i in range(6, -1, -1)]  # 7 days: 6 days ago .. today
    labels = [d.strftime("%d %b") for d in days]
    date_strs = [d.isoformat() for d in days]
    coconut_totals = [0] * 7
    truck_counts = [0] * 7
    conn = _get_db_conn()
    try:
        # Coconut: sum(total) per (year, month, day) for these 7 days
        for i, d in enumerate(days):
            cur = conn.execute(
                """SELECT COALESCE(SUM(total), 0) AS s FROM coconut_count_history
                   WHERE year = ? AND month = ? AND day = ?""",
                (d.year, d.month, d.day),
            )
            row = cur.fetchone()
            coconut_totals[i] = int(row["s"] or 0)
        # Truck driver_snap count per date(created_at)
        for i, d in enumerate(days):
            cur = conn.execute(
                """SELECT COUNT(*) AS n FROM truck_verification_images
                   WHERE capture_type = 'driver_snap' AND date(created_at) = ?""",
                (date_strs[i],),
            )
            row = cur.fetchone()
            truck_counts[i] = int(row["n"] or 0)
        return {"labels": labels, "coconut_totals": coconut_totals, "truck_counts": truck_counts}
    finally:
        conn.close()


@app.route("/api/dashboard/production_last_7_days")
def api_dashboard_production_last_7_days():
    """Return last 7 days daily coconut totals and truck verification counts for dashboard chart."""
    try:
        return jsonify(_get_production_last_7_days())
    except Exception as e:
        print(f"[api] dashboard production_last_7_days error: {e}")
        today = date.today()
        days = [today - timedelta(days=i) for i in range(6, -1, -1)]
        return jsonify({
            "labels": [d.strftime("%d %b") for d in days],
            "coconut_totals": [0] * 7,
            "truck_counts": [0] * 7,
        })


@app.route("/api/operations_efficiency")
def api_operations_efficiency():
    """Return operations efficiency percentages from DB for System Performance Overview chart."""
    try:
        data = _get_operations_efficiency_from_db()
        return jsonify(data)
    except Exception as e:
        print(f"[api] operations_efficiency error: {e}")
        return jsonify({
            "coconut_count": 0,
            "driver_capture": 0,
            "weight_check": 0,
            "license_plate": 0,
        })


@app.route("/api/system_events_history")
def api_system_events_history():
    """Return system events from DB for admin history. Query params: from_date (YYYY-MM-DD), to_date (YYYY-MM-DD), event_type, limit (default 500)."""
    from_date = request.args.get("from_date")
    to_date = request.args.get("to_date")
    event_type = request.args.get("event_type")
    limit = request.args.get("limit", type=int) or 500
    try:
        rows = _get_system_events_history_from_db(from_date=from_date, to_date=to_date, event_type=event_type, limit=limit)
        return jsonify({"events": rows})
    except Exception as e:
        print(f"[api] system_events_history error: {e}")
        return jsonify({"events": []})


def _build_system_report_data(from_date=None, to_date=None, period_label=None):
    """Gather current system state and recent events for report download. Returns dict.
    from_date, to_date: YYYY-MM-DD to filter events; if both None, last 500 events.
    period_label: e.g. 'Today 2025-02-27', 'February 2025', 'Year 2025' for report title/filename."""
    load = get_system_load()
    with coconut_detection_lock:
        coco_enabled = coconut_detection_enabled
        coco_started_at = coconut_detection_started_at
    with truck_detection_lock:
        truck_enabled = truck_detection_enabled
        truck_started_at = truck_detection_started_at
    uptime_sec = (_app_started_at and (datetime.now() - _app_started_at).total_seconds()) or 0
    coco_start = int(_get_system_stat("coconut_model_start_count", "0"))
    coco_stop = int(_get_system_stat("coconut_model_stop_count", "0"))
    truck_start = int(_get_system_stat("truck_model_start_count", "0"))
    truck_stop = int(_get_system_stat("truck_model_stop_count", "0"))
    events = _get_system_events_history_from_db(from_date=from_date, to_date=to_date, limit=2000)
    last_shutdown = None
    try:
        _instance_dir = os.path.join(os.path.dirname(_db_path), "last_shutdown.txt")
        if os.path.isfile(_instance_dir):
            with open(_instance_dir, "r") as f:
                last_shutdown = f.read().strip() or None
    except Exception:
        pass
    return {
        "generated_at": datetime.now().isoformat(),
        "period_label": period_label or "Full history",
        "project": {"name": "SilverMill", "version": "v2.1.0", "build": "2024.12"},
        "system": {"platform": sys.platform, "python": sys.version.split()[0]},
        "load": load,
        "uptime_seconds": int(uptime_sec),
        "app_started_at": _app_started_at.isoformat() if _app_started_at else None,
        "last_shutdown_at": last_shutdown,
        "coconut_model": {"start_count": coco_start, "stop_count": coco_stop, "enabled": coco_enabled, "started_at": coco_started_at.isoformat() if coco_started_at else None},
        "truck_model": {"start_count": truck_start, "stop_count": truck_stop, "enabled": truck_enabled, "started_at": truck_started_at.isoformat() if truck_started_at else None},
        "events": events,
    }


def _report_period_from_request():
    """Parse period=today|monthly|yearly and optional month, year. Returns (from_date, to_date, period_label) for report."""
    period = (request.args.get("period") or "today").strip().lower()
    today = datetime.now().date()
    from_date = to_date = None
    period_label = "Full history"
    if period == "today":
        from_date = to_date = today.isoformat()
        period_label = "Today %s" % from_date
    elif period == "monthly":
        try:
            y = int(request.args.get("year") or today.year)
            m = int(request.args.get("month") or today.month)
        except (TypeError, ValueError):
            y, m = today.year, today.month
        m = max(1, min(12, m))
        from_date = "%04d-%02d-01" % (y, m)
        if m == 12:
            to_date = "%04d-12-31" % y
        else:
            from datetime import date as date_cls, timedelta
            last_day = date_cls(y, m + 1, 1) - timedelta(days=1)
            to_date = last_day.isoformat()
        month_names = ("", "January", "February", "March", "April", "May", "June",
                      "July", "August", "September", "October", "November", "December")
        period_label = "%s %s" % (month_names[m], y)
    elif period == "yearly":
        try:
            y = int(request.args.get("year") or today.year)
        except (TypeError, ValueError):
            y = today.year
        from_date = "%04d-01-01" % y
        to_date = "%04d-12-31" % y
        period_label = "Year %s" % y
    return from_date, to_date, period_label


@app.route("/download_system_report")
def download_system_report():
    """Admin: download system report as CSV or PDF.
    Query: format=csv|pdf, period=today|monthly|yearly, month=1-12 (for monthly), year=YYYY (for monthly/yearly)."""
    fmt = (request.args.get("format") or "csv").strip().lower()
    if fmt not in ("csv", "pdf"):
        fmt = "csv"
    from_date, to_date, period_label = _report_period_from_request()
    try:
        data = _build_system_report_data(from_date=from_date, to_date=to_date, period_label=period_label)
    except Exception as e:
        print(f"[download_system_report] build error: {e}")
        return jsonify({"error": "Failed to build report"}), 500
    safe_label = "".join(c if c.isalnum() or c in "-_" else "_" for c in period_label)
    if fmt == "csv":
        buffer = io.BytesIO()
        writer = io.TextIOWrapper(buffer, encoding="utf-8-sig", newline="")
        writer.write("SilverMill System Report - %s\r\n" % data["period_label"])
        writer.write("Generated,%s\r\n" % data["generated_at"])
        writer.write("\r\n")
        writer.write("Project,%s\r\n" % data["project"]["name"])
        writer.write("Version,%s\r\n" % data["project"]["version"])
        writer.write("Build,%s\r\n" % data["project"]["build"])
        writer.write("Platform,%s\r\n" % data["system"]["platform"])
        writer.write("Python,%s\r\n" % data["system"]["python"])
        writer.write("Uptime (seconds),%s\r\n" % data["uptime_seconds"])
        writer.write("App started at,%s\r\n" % (data["app_started_at"] or ""))
        writer.write("Last shutdown at,%s\r\n" % (data["last_shutdown_at"] or ""))
        writer.write("\r\n")
        writer.write("Coconut model start count (times on),%s\r\n" % data["coconut_model"]["start_count"])
        writer.write("Coconut model stop count (times off),%s\r\n" % data["coconut_model"]["stop_count"])
        writer.write("Coconut model status,%s\r\n" % ("On" if data["coconut_model"]["enabled"] else "Off"))
        writer.write("Coconut model started at,%s\r\n" % (data["coconut_model"]["started_at"] or ""))
        writer.write("\r\n")
        writer.write("Truck model start count (times on),%s\r\n" % data["truck_model"]["start_count"])
        writer.write("Truck model stop count (times off),%s\r\n" % data["truck_model"]["stop_count"])
        writer.write("Truck model status,%s\r\n" % ("On" if data["truck_model"]["enabled"] else "Off"))
        writer.write("Truck model started at,%s\r\n" % (data["truck_model"]["started_at"] or ""))
        writer.write("\r\n")
        writer.write("CPU %%,%s\r\n" % (data["load"].get("cpu_pct") if data["load"].get("cpu_pct") is not None else ""))
        writer.write("RAM %%,%s\r\n" % (data["load"].get("ram_pct") if data["load"].get("ram_pct") is not None else ""))
        writer.write("GPU,,%s\r\n" % (data["load"].get("gpu_util") or data["load"].get("gpu_mem_used") or ""))
        writer.write("\r\n")
        writer.write("Event Type,Message,Level,Created At\r\n")
        for ev in data["events"]:
            msg = (ev.get("message") or "").replace("\r", " ").replace("\n", " ")
            writer.write("%s,%s,%s,%s\r\n" % (ev.get("event_type", ""), msg, ev.get("level", ""), ev.get("time", "")))
        writer.flush()
        buffer.seek(0)
        filename = "system_report_%s_%s.csv" % (safe_label.replace(" ", "_"), datetime.now().strftime("%Y-%m-%d_%H%M"))
        return send_file(buffer, mimetype="text/csv", as_attachment=True, download_name=filename)
    else:
        if not REPORTLAB_AVAILABLE:
            return jsonify({"error": "PDF requires reportlab. pip install reportlab"}), 500
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch, leftMargin=0.5*inch, rightMargin=0.5*inch)
        styles = getSampleStyleSheet()
        elements = []
        elements.append(Paragraph("SilverMill System Report", styles["Title"]))
        elements.append(Spacer(1, 0.15*inch))
        # Table: Report info
        tbl_report = Table([
            ["Report period", data["period_label"]],
            ["Generated", data["generated_at"]],
        ], colWidths=[2*inch, 4*inch])
        tbl_report.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#e5e7eb")),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(tbl_report)
        elements.append(Spacer(1, 0.25*inch))
        # Table: Project & System
        tbl_project = Table([
            ["Item", "Value"],
            ["Project", data["project"]["name"]],
            ["Version", data["project"]["version"]],
            ["Build", data["project"]["build"]],
            ["Platform", data["system"]["platform"]],
            ["Python", data["system"]["python"]],
        ], colWidths=[2*inch, 4*inch])
        tbl_project.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6a11cb")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#f3f4f6")),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(Paragraph("Project &amp; System", styles["Heading2"]))
        elements.append(Spacer(1, 0.08*inch))
        elements.append(tbl_project)
        elements.append(Spacer(1, 0.2*inch))
        # Table: Server status
        tbl_server = Table([
            ["Field", "Value"],
            ["Uptime (seconds)", str(data["uptime_seconds"])],
            ["App started at", data["app_started_at"] or "—"],
            ["Last shutdown at", data["last_shutdown_at"] or "—"],
        ], colWidths=[2*inch, 4*inch])
        tbl_server.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#f3f4f6")),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(Paragraph("Server Status", styles["Heading2"]))
        elements.append(Spacer(1, 0.08*inch))
        elements.append(tbl_server)
        elements.append(Spacer(1, 0.2*inch))
        # Table: Coconut model
        coco = data["coconut_model"]
        tbl_coco = Table([
            ["Field", "Value"],
            ["Times turned ON (lifetime)", str(coco["start_count"])],
            ["Times turned OFF (lifetime)", str(coco.get("stop_count", 0))],
            ["Status", "On" if coco["enabled"] else "Off"],
            ["Started at", coco["started_at"] or "—"],
        ], colWidths=[2*inch, 4*inch])
        tbl_coco.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#065f46")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#f3f4f6")),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(Paragraph("Coconut Model", styles["Heading2"]))
        elements.append(Spacer(1, 0.08*inch))
        elements.append(tbl_coco)
        elements.append(Spacer(1, 0.2*inch))
        # Table: Truck model
        truck = data["truck_model"]
        tbl_truck = Table([
            ["Field", "Value"],
            ["Times turned ON (lifetime)", str(truck["start_count"])],
            ["Times turned OFF (lifetime)", str(truck["stop_count"])],
            ["Status", "On" if truck["enabled"] else "Off"],
            ["Started at", truck["started_at"] or "—"],
        ], colWidths=[2*inch, 4*inch])
        tbl_truck.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#f3f4f6")),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(Paragraph("Truck Model (LNPR / Driver)", styles["Heading2"]))
        elements.append(Spacer(1, 0.08*inch))
        elements.append(tbl_truck)
        elements.append(Spacer(1, 0.2*inch))
        # Table: Resource (CPU / RAM / GPU)
        load = data["load"]
        cpu_val = str(load.get("cpu_pct")) + "%" if load.get("cpu_pct") is not None else "—"
        ram_val = str(load.get("ram_pct")) + "%" if load.get("ram_pct") is not None else "—"
        gpu_parts = []
        if load.get("gpu_util") is not None:
            gpu_parts.append(str(load["gpu_util"]))
        if load.get("gpu_mem_used") is not None:
            if load.get("gpu_mem_total") is not None:
                gpu_parts.append("%s/%s MB" % (load["gpu_mem_used"], load["gpu_mem_total"]))
            else:
                gpu_parts.append("%s MB" % load["gpu_mem_used"])
        gpu_val = " ".join(gpu_parts) if gpu_parts else "—"
        tbl_res = Table([
            ["Resource", "Value"],
            ["CPU", cpu_val],
            ["RAM", ram_val],
            ["GPU", gpu_val],
        ], colWidths=[2*inch, 4*inch])
        tbl_res.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4b5563")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#f3f4f6")),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(Paragraph("Resource Usage", styles["Heading2"]))
        elements.append(Spacer(1, 0.08*inch))
        elements.append(tbl_res)
        elements.append(Spacer(1, 0.25*inch))
        # Table: System events
        elements.append(Paragraph("System Events (%s)" % data["period_label"], styles["Heading2"]))
        elements.append(Spacer(1, 0.08*inch))
        ev_data = [["Time", "Type", "Level", "Message"]]
        for ev in data["events"][:200]:
            ev_data.append([(ev.get("time") or "")[:19], ev.get("event_type", ""), ev.get("level", ""), (ev.get("message") or "")[:80]])
        t_ev = Table(ev_data, colWidths=[inch*1.8, inch*1.2, 0.5*inch, inch*3.5])
        t_ev.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(t_ev)
        doc.build(elements)
        buffer.seek(0)
        filename = "system_report_%s_%s.pdf" % (safe_label.replace(" ", "_"), datetime.now().strftime("%Y-%m-%d_%H%M"))
        return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=filename)


@app.route("/api/coconut_count_history")
def api_coconut_count_history():
    """Return count history rows from SQLite for the Count table (weight_check page). Supports optional year, month, day, time (shift1/shift2/shift3)."""
    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)
    day = request.args.get("day", type=int)
    time_filter = request.args.get("time")  # 'shift1', 'shift2', 'shift3' or None
    if time_filter and time_filter not in ("shift1", "shift2", "shift3"):
        time_filter = None
    try:
        rows = _get_coconut_history_from_db(year=year, month=month, day=day, time_filter=time_filter)
        return jsonify({"rows": rows})
    except Exception as e:
        print(f"[api] coconut_count_history error: {e}")
        return jsonify({"rows": []})


@app.route("/download_coconut_report_pdf", methods=["GET", "POST"])
def download_coconut_report_pdf():
    """Generate and download Coconut Count report PDF with SilverMill and partner logos."""
    if not REPORTLAB_AVAILABLE:
        return jsonify({"error": "reportlab not installed. pip install reportlab"}), 500
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.4*inch, bottomMargin=0.5*inch,
                               leftMargin=0.5*inch, rightMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CoconutTitle', parent=styles['Heading1'],
            fontSize=18, textColor=colors.HexColor('#1e40af'), alignment=1, spaceAfter=6
        )
        section_style = ParagraphStyle(
            'Section', parent=styles['Normal'],
            fontSize=12, textColor=colors.HexColor('#374151'), spaceAfter=10, spaceBefore=14
        )

        # Logos side by side in header
        static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'images')
        logo_path = os.path.join(static_dir, 'logo.png')
        logo1_path = os.path.join(static_dir, 'logo1.png')
        logo_w, logo_h = 1.6*inch, 0.55*inch
        logo_cells = []
        if os.path.isfile(logo_path):
            logo_cells.append(Image(logo_path, width=logo_w, height=logo_h))
        else:
            logo_cells.append(Paragraph("<i>SilverMill</i>", styles['Normal']))
        if os.path.isfile(logo1_path):
            logo_cells.append(Image(logo1_path, width=logo_w, height=logo_h))
        else:
            logo_cells.append(Paragraph("", styles['Normal']))
        if len(logo_cells) >= 2:
            logo_table = Table([logo_cells], colWidths=[2.5*inch, 2.5*inch])
            logo_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(logo_table)
        elif len(logo_cells) == 1:
            elements.append(logo_cells[0])
        elements.append(Spacer(1, 0.15*inch))

        report_date = datetime.now().strftime('%B %d, %Y')
        elements.append(Paragraph(f"Coconut Count Report", title_style))
        elements.append(Paragraph(f"Generated on {report_date}", section_style))
        elements.append(Spacer(1, 0.2*inch))

        with coconut_counts_lock:
            counts = dict(coconut_counts)
        with coconut_detection_lock:
            enabled = coconut_detection_enabled
            started_at = coconut_detection_started_at
        history = _get_coconut_history_from_db()

        c0 = counts.get(0, 0) + counts.get(1, 0)
        c1 = counts.get(2, 0) + counts.get(3, 0)
        c2 = counts.get(4, 0) + counts.get(5, 0)
        belt1, belt2, belt3 = c0, c1, c2
        total = belt1 + belt2 + belt3

        duration_str = "—"
        if enabled and started_at is not None:
            delta = datetime.now() - started_at
            total_sec = int(delta.total_seconds())
            h, r = divmod(total_sec, 3600)
            m, s = divmod(r, 60)
            duration_str = f"{h}h {m}m {s}s"

        elements.append(Paragraph("Summary", section_style))
        info_data = [
            ["Today's count (live)", str(total)],
            ["Belt 1 (Cam 1+2)", str(belt1)],
            ["Belt 2 (Cam 3+4)", str(belt2)],
            ["Belt 3 (Cam 5+6)", str(belt3)],
            ["Model running", "Yes" if enabled else "No"],
            ["Duration (this session)", duration_str],
        ]
        table = Table(info_data, colWidths=[2.6*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.25*inch))

        # Count history table (all rows)
        elements.append(Paragraph("Count History", section_style))
        col_w = [0.85*inch, 0.5*inch, 0.7*inch, 0.5*inch, 0.6*inch, 0.65*inch, 0.55*inch, 0.55*inch, 0.65*inch]
        table_data = [
            ["Date", "Time", "Total", "Batches", "Avg/Batch", "Belt", "Cameras", "Accuracy", "Status"]
        ]
        for row in history[-150:]:  # last 150 rows (newest last in PDF)
            table_data.append([
                row.get("date", ""),
                row.get("time", ""),
                str(row.get("total", 0)),
                str(row.get("batches", "")),
                str(row.get("avgBatch", "")),
                row.get("conveyorBelt", ""),
                row.get("camera", "Cam 1-6"),
                row.get("accuracy", "99.2%"),
                row.get("status", "Completed"),
            ])
        if len(table_data) == 1:
            table_data.append(["—", "—", "0", "—", "—", "—", "—", "—", "No data yet"])
        tbl = Table(table_data, colWidths=col_w, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fdf4')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('ALIGN', (2, 0), (4, -1), 'CENTER'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
        elements.append(tbl)

        doc.build(elements)
        buffer.seek(0)
        filename = f"coconut_report_{datetime.now().strftime('%Y-%m-%d_%H%M')}.pdf"
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------------- System load stats (terminal: CPU, GPU, balance) ----------------
def _get_gpu_stats():
    """Return GPU utilization % and memory (used/total MB). Uses nvidia-smi or torch.cuda."""
    gpu_util = None
    gpu_mem_used = None
    gpu_mem_total = None
    try:
        creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0) if sys.platform == "win32" else 0
        out = subprocess.run(
            ["nvidia-smi", "--query-gpu=utilization.gpu,memory.used,memory.total", "--format=csv,noheader,nounits"],
            capture_output=True, text=True, timeout=2, creationflags=creationflags
        )
        if out.returncode == 0 and out.stdout.strip():
            line = out.stdout.strip().split("\n")[0]
            parts = [p.strip().split()[0] if p.strip() else "" for p in line.split(",")]
            if len(parts) >= 3:
                gpu_util, gpu_mem_used, gpu_mem_total = parts[0], parts[1], parts[2]
    except Exception:
        pass
    if gpu_util is None and gpu_mem_used is None:
        try:
            import torch
            if torch.cuda.is_available():
                gpu_mem_used = str(torch.cuda.memory_allocated(0) // (1024 * 1024))
                gpu_mem_total = str(torch.cuda.get_device_properties(0).total_memory // (1024 * 1024))
                gpu_util = "N/A"
        except Exception:
            pass
    return gpu_util, gpu_mem_used, gpu_mem_total


def get_system_load():
    """Return current system load as dict for API and terminal: CPU, RAM, GPU, Detection, Cameras."""
    out = {
        "cpu_pct": None,
        "ram_pct": None,
        "ram_used_m": None,
        "ram_total_m": None,
        "gpu_util": None,
        "gpu_mem_used": None,
        "gpu_mem_total": None,
        "detection": "OFF",
        "cameras_active": 0,
        "cameras_total": 6,
    }
    if PSUTIL_AVAILABLE:
        out["cpu_pct"] = round(psutil.cpu_percent(interval=None), 1)
        mem = psutil.virtual_memory()
        out["ram_pct"] = round(mem.percent, 1)
        out["ram_used_m"] = mem.used // (1024 * 1024)
        out["ram_total_m"] = mem.total // (1024 * 1024)
    gpu_util, gpu_mem_used, gpu_mem_total = _get_gpu_stats()
    out["gpu_util"] = gpu_util
    out["gpu_mem_used"] = int(gpu_mem_used) if gpu_mem_used is not None else None
    out["gpu_mem_total"] = int(gpu_mem_total) if gpu_mem_total is not None else None
    with coconut_detection_lock:
        out["detection"] = "ON" if coconut_detection_enabled else "OFF"
    coconut_cams = list(range(6))
    active = 0
    for c in coconut_cams:
        if c not in _camera_frame_buffers:
            continue
        with _camera_buffer_locks[c]:
            data = _camera_frame_buffers.get(c)
        if data is not None and data[0] is not None:
            active += 1
    out["cameras_active"] = active
    out["cameras_total"] = len(coconut_cams)
    return out


@app.route("/api/system_load")
def api_system_load():
    """Return [Load] stats for dashboard right sidebar: CPU, RAM, GPU, Detection, Cameras."""
    return jsonify(get_system_load())


@app.route("/api/dashboard_info")
def api_dashboard_info():
    """Return full dashboard payload: system load, detection, uptime, project info, DB status, cameras, system events."""
    global _app_started_at
    load = get_system_load()
    with coconut_detection_lock:
        detection_enabled = coconut_detection_enabled
        detection_started_at = coconut_detection_started_at
    uptime_sec = (datetime.now() - _app_started_at).total_seconds() if _app_started_at else 0
    project = {
        "name": "SilverMill",
        "version": "v2.1.0",
        "build": "2024.12",
        "status": "Online",
    }
    db_path = _db_path
    try:
        conn = _get_db_conn()
        conn.close()
        db_ok = True
    except Exception:
        db_ok = False
    cameras_total = len(camera_urls)
    with _system_events_lock:
        events = list(_system_events[-50:])
    if not events:
        _log_system_event("Application started (server boot). System is running.", "success", event_type="app_start")
        with _system_events_lock:
            events = list(_system_events[-50:])
    try:
        coconut_model_start_count = int(_get_system_stat("coconut_model_start_count", "0"))
    except Exception:
        coconut_model_start_count = 0
    try:
        coconut_model_stop_count = int(_get_system_stat("coconut_model_stop_count", "0"))
    except Exception:
        coconut_model_stop_count = 0
    try:
        truck_model_start_count = int(_get_system_stat("truck_model_start_count", "0"))
    except Exception:
        truck_model_start_count = 0
    try:
        truck_model_stop_count = int(_get_system_stat("truck_model_stop_count", "0"))
    except Exception:
        truck_model_stop_count = 0
    with truck_detection_lock:
        truck_enabled = truck_detection_enabled
        truck_started_at = truck_detection_started_at
    last_shutdown_at = None
    _instance_dir = os.path.join(os.path.dirname(_db_path), "last_shutdown.txt")
    try:
        if os.path.isfile(_instance_dir):
            with open(_instance_dir, "r") as f:
                last_shutdown_at = f.read().strip() or None
    except Exception:
        pass
    return jsonify({
        "load": load,
        "detection": {"enabled": detection_enabled, "started_at": detection_started_at.isoformat() if detection_started_at else None},
        "truck_detection": {"enabled": truck_enabled, "started_at": truck_started_at.isoformat() if truck_started_at else None},
        "uptime_seconds": int(uptime_sec),
        "app_started_at": _app_started_at.isoformat() if _app_started_at else None,
        "coconut_model_start_count": coconut_model_start_count,
        "coconut_model_stop_count": coconut_model_stop_count,
        "truck_model_start_count": truck_model_start_count,
        "truck_model_stop_count": truck_model_stop_count,
        "last_shutdown_at": last_shutdown_at,
        "project": project,
        "database": {"ok": db_ok},
        "cameras_total": cameras_total,
        "cameras_active": load.get("cameras_active", 0),
        "system_events": events,
        "system": {"platform": sys.platform, "python": sys.version.split()[0]},
    })


def _system_stats_loop():
    """Daemon thread: print CPU, GPU, RAM and load balance to terminal every 2.5s."""
    interval = 2.5
    while True:
        time.sleep(interval)
        load = get_system_load()
        parts = []
        if load["cpu_pct"] is not None:
            parts.append(f"CPU {load['cpu_pct']:.0f}%")
            parts.append(f"RAM {load['ram_pct']:.0f}% ({load['ram_used_m']}M/{load['ram_total_m']}M)")
        else:
            parts.append("CPU N/A (install psutil)")
        if load["gpu_util"] is not None or load["gpu_mem_used"] is not None:
            gpu_str = "GPU"
            if load["gpu_util"] is not None:
                gpu_str += f" {load['gpu_util']}"
            if load["gpu_mem_used"] is not None and load["gpu_mem_total"] is not None:
                gpu_str += f" Mem {load['gpu_mem_used']}/{load['gpu_mem_total']}MB"
            elif load["gpu_mem_used"] is not None:
                gpu_str += f" Mem {load['gpu_mem_used']}MB"
            parts.append(gpu_str)
        parts.append(f"Detection {load['detection']}")
        parts.append(f"Cameras {load['cameras_active']}/{load['cameras_total']}")
        line = " | ".join(parts)
        try:
            print(f"[Load] {line}", flush=True)
        except Exception:
            pass


def start_system_stats_thread():
    """Start background thread that prints CPU/GPU/RAM and load to terminal."""
    t = threading.Thread(target=_system_stats_loop, daemon=True)
    t.start()


# ---------------- Shutdown: record last exit for System page ----------------
def _write_last_shutdown():
    """Write timestamp to instance/last_shutdown.txt and log shutdown event to DB for history."""
    try:
        now = datetime.now()
        try:
            _insert_system_event_to_db("Application shutdown.", level="info", event_type="app_shutdown", created_at=now)
        except Exception:
            pass
        d = os.path.dirname(_db_path)
        os.makedirs(d, exist_ok=True)
        path = os.path.join(d, "last_shutdown.txt")
        with open(path, "w") as f:
            f.write(now.isoformat())
    except Exception:
        pass


atexit.register(_write_last_shutdown)


# ---------------- Run App ----------------
def _start_model_preload():
    """Preload coconut and ALPR models in background so detection is smooth on first use."""
    def _run():
        time.sleep(2)
        _preload_coconut_model()
        _preload_truck_alpr_model()
    threading.Thread(target=_run, daemon=True).start()


def start_coconut_history_thread():
    """Start background thread that appends count snapshot every 30s when detection is on."""
    t = threading.Thread(target=_coconut_history_loop, daemon=True)
    t.start()


def start_cache_clear_thread():
    """Start background thread that clears all caches every 6 hours for system performance."""
    t = threading.Thread(target=_cache_clear_loop, daemon=True)
    t.start()


def run_startup_tasks():
    """Run all background threads and preloads. Used by desktop launcher and normal run."""
    start_system_stats_thread()
    start_coconut_history_thread()
    start_cache_clear_thread()
    _start_model_preload()


if __name__ == "__main__":
    run_startup_tasks()
    app.run(host="0.0.0.0", port=8080, debug=True)



